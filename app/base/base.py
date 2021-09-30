import psycopg2
from openpyxl import load_workbook
import config


class Base:
    def __init__(self):
        self.conn = psycopg2.connect(dbname=config.base_name, user=config.base_user,
                                     password=config.base_password, host=config.base_host)
        self.cursor = self.conn.cursor()

    def create_tablets(self):
        with open('CREATE_TABLES_SCRIPT.sql', 'r') as file:
            sql_script = file.read()
        self.cursor.execute(sql_script)

    def insert_specialist_status_data_in_db(self, status_data):
        for row in status_data:
            sql_script = f"INSERT INTO specialist_statuses (status_name) VALUES ('{row}')"
            self.cursor.execute(sql_script)
            print(f'Значение {row} добавлено!')

    def insert_specialists_data_in_db(self, specialists_data):
        for row in specialists_data:
            sql_script = f"INSERT INTO specialists (specialist_name, fk_status_specialist) VALUES (" \
                         f"'{row['name']}'," \
                         f"(SELECT status_id FROM specialist_statuses WHERE status_name = '{row['status']}'))"
            self.cursor.execute(sql_script)
            print(f"Сотрудник {row['name']} добавлен!")

    def insert_group_services_data_in_db(self, group_services_data):
        for row in group_services_data:
            sql_script = f"INSERT INTO group_services (group_name) VALUES ('{row}')"
            self.cursor.execute(sql_script)
            print(f'Группа услуг {row} добавлена')

    def insert_tags_services_data_in_db(self, tags_services_data):
        for row in tags_services_data:
            sql_script = f"INSERT INTO tags_services (tag_name) VALUES ('{row}')"
            self.cursor.execute(sql_script)
            print(f'Тег {row} добавлен!')

    def insert_services_data_in_db(self, services_data):
        for row in services_data:
            sql_script = f"INSERT INTO services (service_name, fk_group_service) VALUES ('{row['name']}', " \
                         f"(SELECT group_id FROM group_services WHERE group_name = '{row['group']}'))"
            self.cursor.execute(sql_script)
            print(f"Услуга {row['name']} добавлена!")

    def insert_service_in_db(self, service):
        sql_script = f"""INSERT INTO services (service_name, fk_group_service) VALUES (
        '{service['service_name']}',
        {service['group-select']}
        )"""
        self.cursor.execute(sql_script)
        print(f"Услуга {service['service_name']} добавлена!")

    def insert_program_services_data_in_db(self, program_services_data):
        for row in program_services_data:
            sql_script = f"INSERT INTO program_services (program_service_name, fk_tag_service) VALUES (" \
                         f"'{row['name']}', " \
                         f"(SELECT tag_id FROM tags_services WHERE tag_name = '{row['tag']}'))"
            self.cursor.execute(sql_script)
            print(f"Программная услуга {row['name']} добавлена!")

    def insert_program_service_in_db(self, program_service):
        sql_script = f"""INSERT INTO program_services (program_service_name, fk_tag_service) VALUES (
        '{program_service['name-program-service']}',
        {program_service['tag-select']}
        )"""
        self.cursor.execute(sql_script)
        print(f"Программная услуга {program_service['name-program-service']} добавлена!")

    def insert_reception_table_in_db(self, reception_data):
        sql_script = f"""INSERT INTO reception_table (date_reception, fk_specialist, fk_service, count_reception)
            VALUES (
            '{reception_data['date_reception']}',
            (SELECT specialist_id FROM specialists WHERE specialist_name = '{reception_data['specialist_name']}'),
            (SELECT program_service_id FROM program_services WHERE program_service_name = '{reception_data['service_name']}'),
            {reception_data['count_reception']}
            );"""
        self.cursor.execute(sql_script)
        print("Запсиь о приёме добавлена!")

    def get_program_services_table(self):
        sql_script = """SELECT program_services.program_service_name, services.service_name, 
            program_services.program_service_id, services.service_id 
            FROM program_services
            LEFT JOIN services_program_services 
            ON services_program_services.fk_program_service = program_services.program_service_id
            LEFT JOIN services 
            ON services_program_services.fk_service = services.service_id
            ORDER BY services.service_name;"""
        self.cursor.execute(sql_script)
        program_services = self.cursor.fetchall()
        return program_services

    def get_services_table(self):
        sql_script = """SELECT services.service_name FROM services;"""
        self.cursor.execute(sql_script)
        return self.cursor.fetchall()

    def get_all_services_table(self):
        sql_script = """SELECT group_services.group_name, service_name, service_id FROM services
            LEFT JOIN group_services ON services.fk_group_service = group_services.group_id
            ORDER BY group_services.group_name, service_name;"""
        self.cursor.execute(sql_script)
        return self.cursor.fetchall()

    def set_services_program_services(self, data):
        sql_script_delete = """DELETE FROM services_program_services;"""
        self.cursor.execute(sql_script_delete)
        for p_service_id, service_name in data.items():
            sql_script = f"""INSERT INTO services_program_services (fk_program_service, fk_service) VALUES (
                {p_service_id}, 
                (SELECT service_id FROM services WHERE service_name = '{service_name}')
            );"""
            self.cursor.execute(sql_script)

    def get_service_info(self, service_id):
        sql_script = f"""SELECT * FROM services WHERE service_id = {service_id}"""
        self.cursor.execute(sql_script)
        return self.cursor.fetchall()

    def get_tags_services(self):
        sql_script = f"""SELECT * FROM tags_services"""
        self.cursor.execute(sql_script)
        return self.cursor.fetchall()

    def get_specialists(self):
        sql_script = """SELECT status_name, specialist_name, rating, specialist_id FROM specialists
            LEFT JOIN specialist_statuses ON specialists.fk_status_specialist = specialist_statuses.status_id
            ORDER BY status_id, specialist_name;"""
        self.cursor.execute(sql_script)
        return self.cursor.fetchall()

    def get_group_services(self):
        sql_script = """SELECT * FROM group_services;"""
        self.cursor.execute(sql_script)
        return self.cursor.fetchall()

    def get_report_services(self, first_date, last_date):
        sql_script = f"""
        SELECT group_services.group_name, service_name, sum(count_reception) FROM reception_table
        LEFT JOIN specialists ON reception_table.fk_specialist = specialist_id
        LEFT JOIN program_services ON reception_table.fk_service = program_service_id
        LEFT JOIN services_program_services ON program_service_id = services_program_services.fk_program_service
        LEFT JOIN services ON services_program_services.fk_service = service_id
        LEFT JOIN group_services ON services.fk_group_service = group_id
        WHERE date_reception >= '{first_date}'
        AND date_reception <= '{last_date}'
        GROUP BY service_name, group_services.group_name
        ORDER BY group_services.group_name;
        """
        self.cursor.execute(sql_script)
        return self.cursor.fetchall()

    def commit_bd(self):
        self.conn.commit()

    def close_bd(self):
        self.conn.close()


class ExcelBase:
    def __init__(self):
        self.wb = load_workbook('base_model.xlsx')

    def parse_status_specialist(self):
        sheet_status_specialist = self.wb['status_specialists']
        statuses_specialists = []
        rows = sheet_status_specialist.max_row
        for i in range(2, rows + 1):
            statuses_specialists.append(sheet_status_specialist.cell(row=i, column=1).value)
        return statuses_specialists

    def parse_specialists(self):
        sheet_specialists = self.wb['specialists']
        specialists = []
        rows = sheet_specialists.max_row
        for i in range(2, rows + 1):
            name = sheet_specialists.cell(row=i, column=1).value
            status = sheet_specialists.cell(row=i, column=2).value
            specialists.append({'name': name, 'status': status})
        return specialists

    def parse_group_services(self):
        sheet_group_services = self.wb['groups']
        groups_services = []
        rows = sheet_group_services.max_row
        for i in range(2, rows + 1):
            groups_services.append(sheet_group_services.cell(row=i, column=1).value)
        return groups_services

    def parse_tags_services(self):
        sheet_group_tags_services = self.wb['tegs']
        tags_services = []
        rows = sheet_group_tags_services.max_row
        for i in range(2, rows + 1):
            tags_services.append(sheet_group_tags_services.cell(row=i, column=1).value)
        return tags_services

    def parse_services(self):
        sheet_services = self.wb['services']
        services = []
        rows = sheet_services.max_row
        for i in range(2, rows + 1):
            name = sheet_services.cell(row=i, column=1).value
            group = sheet_services.cell(row=i, column=2).value
            services.append({'name': name, 'group': group})
        return services

    def parse_program_services(self):
        sheet_program_services = self.wb['program_services']
        program_services = []
        rows = sheet_program_services.max_row
        for i in range(2, rows + 1):
            name = sheet_program_services.cell(row=i, column=1).value
            tag = sheet_program_services.cell(row=i, column=2).value
            program_services.append({'name': name, 'tag': tag})
        return program_services


if __name__ == '__main__':
    bd = Base()
    excel_data = ExcelBase()

    # Создание таблиц БД
    # bd.create_tablets()

    # Заполнение таблицы статусы специалистов
    # status_data = excel_data.parse_status_specialist()
    # bd.insert_specialist_status_data_in_db(status_data)

    # Заполнение таблицы списка специалистов
    # specialists_data = excel_data.parse_specialists()
    # bd.insert_specialists_data_in_db(specialists_data)

    # Заполнение таблицы списка групп услуг
    # group_services_data = excel_data.parse_group_services()
    # bd.insert_group_services_data_in_db(group_services_data)

    # Заполнение таблицы списка тегов услуг
    # tags_services_data = excel_data.parse_tags_services()
    # bd.insert_tags_services_data_in_db(tags_services_data)

    # Заполнение таблицы списка услуг
    # services_data = excel_data.parse_services()
    # bd.insert_services_data_in_db(services_data)

    # Заполнение таблицы списка программных услуг
    # program_services_data = excel_data.parse_program_services()
    # bd.insert_program_services_data_in_db(program_services_data)

    bd.commit_bd()
    bd.close_bd()
