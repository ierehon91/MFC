import os
import openpyxl
import datetime

from base.base import Base


def parse_data(files):
    data = {'reception': [], 'gibrit': [], 'EO': []}
    for file in files:
        wb = openpyxl.load_workbook(directory + '\\' + file)
        reception_data = parse_reception(wb)
        data['reception'] += reception_data
    return data


def insert_data(data):
    db_data = Base()
    for row in data['reception']:
        db_data.insert_reception_table_in_db(row)
    db_data.commit_bd()
    db_data.close_bd()


def parse_reception(wb):
    sheet_reception = wb['Приём']
    reception = []
    print(sheet_reception.max_row)
    if sheet_reception.max_row > 1:
        for i in range(1, sheet_reception.max_row + 1):
            date_reception = sheet_reception.cell(row=i, column=1).value
            specialist_name = sheet_reception.cell(row=i, column=2).value
            service_name = sheet_reception.cell(row=i, column=3).value
            count_reception = sheet_reception.cell(row=i, column=4).value
            date_reception = date_reception.split('.')
            date_reception = datetime.date(int(date_reception[2]), int(date_reception[1]), int(date_reception[0])).isoformat()
            reception.append({'date_reception': date_reception,
                              'specialist_name': specialist_name,
                              'service_name': service_name,
                              'count_reception': count_reception
                              })
        return reception
    else:
        return []


def parse_gibrit(wb):
    sheet_gibrit = wb['Гибрит']
    gibrit = []
    for i in range(1, sheet_gibrit.max_row + 1):
        status_pay = sheet_gibrit.cell(row=i, column=1).value
        suip = sheet_gibrit.cell(row=i, column=2).value
        date_pay = sheet_gibrit.cell(row=i, column=3).value
        specialist_name_pay = sheet_gibrit.cell(row=i, column=4).value
        inn_pay = sheet_gibrit.cell(row=i, column=5).value
        service_pay = sheet_gibrit.cell(row=i, column=6).value
        organization_pay = sheet_gibrit.cell(row=i, column=7).value
        podrazdelenie_pay = sheet_gibrit.cell(row=i, column=8).value
        window_pay = sheet_gibrit.cell(row=i, column=9).value
        comission_pay = sheet_gibrit.cell(row=i, column=10).value
        total_pay = sheet_gibrit.cell(row=i, column=11).value
        gibrit.append({'status_pay': status_pay,
                       'suip': suip,
                       'date_pay': date_pay,
                       'specialist_name_pay': specialist_name_pay,
                       'inn_pay': inn_pay,
                       'service_pay': service_pay,
                       'organization_pay': organization_pay,
                       'podrazdelenie_pay': podrazdelenie_pay,
                       'window_pay': window_pay,
                       'comission_pay': comission_pay,
                       'total_pay': total_pay
                       })
    return gibrit


if __name__ == '__main__':
    directory = '..\\temp\\day_reports'
    files = os.listdir(directory)
    data = parse_data(files)
    insert_data(data)
