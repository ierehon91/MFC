import psycopg2
from openpyxl import load_workbook
import app.mfc_config as config


class ExcelBase:
    """Класс парсинга данных из excel таблицы base_model"""
    def __init__(self):
        self.wb = load_workbook('base_model.xlsx')

    def parse_status_specialist(self):
        """Получение данных со статусами специалистов"""
        sheet_status_specialist = self.wb['status_specialists']
        statuses_specialists = []
        rows = sheet_status_specialist.max_row
        for i in range(2, rows + 1):
            statuses_specialists.append(sheet_status_specialist.cell(row=i, column=1).value)
        return statuses_specialists

    def parse_specialists(self):
        """Получение данных о специалистах"""
        sheet_specialists = self.wb['specialists']
        specialists = []
        rows = sheet_specialists.max_row
        for i in range(2, rows + 1):
            name = sheet_specialists.cell(row=i, column=1).value
            status = sheet_specialists.cell(row=i, column=2).value
            specialists.append({'name': name, 'status': status})
        return specialists

    def parse_group_services(self):
        """Получение данных о группах услуг"""
        sheet_group_services = self.wb['groups']
        groups_services = []
        rows = sheet_group_services.max_row
        for i in range(2, rows + 1):
            groups_services.append(sheet_group_services.cell(row=i, column=1).value)
        return groups_services

    def parse_tags_services(self):
        """Получение данных о тегах услуг"""
        sheet_group_tags_services = self.wb['tegs']
        tags_services = []
        rows = sheet_group_tags_services.max_row
        for i in range(2, rows + 1):
            tags_services.append(sheet_group_tags_services.cell(row=i, column=1).value)
        return tags_services

    def parse_services(self):
        """Получение данных об услугах"""
        sheet_services = self.wb['services']
        services = []
        rows = sheet_services.max_row
        for i in range(2, rows + 1):
            name = sheet_services.cell(row=i, column=1).value
            group = sheet_services.cell(row=i, column=2).value
            services.append({'name': name, 'group': group})
        return services

    def parse_program_services(self):
        """Получение данных о программных услугах"""
        sheet_program_services = self.wb['program_services']
        program_services = []
        rows = sheet_program_services.max_row
        for i in range(2, rows + 1):
            name = sheet_program_services.cell(row=i, column=1).value
            tag = sheet_program_services.cell(row=i, column=2).value
            program_services.append({'name': name, 'tag': tag})
        return program_services


class UpdateBase:
    def __init__(self):
        self.conn = psycopg2.connect(dbname=config.base_name, user=config.base_user,
                                     password=config.base_password, host=config.base_host)
        self.cursor = self.conn.cursor()

    def create_tablets(self):
        """Парсит sql скрипты создания таблиц и выполняет их"""
        with open('CREATE_TABLES_SCRIPT.sql', 'r') as file:
            sql_script = file.read()
        self.cursor.execute(sql_script)

    def insert_specialist_status_data_in_db(self, status_data):
        """Добавление статуса специалиста"""
        for row in status_data:
            sql_script = f"INSERT INTO specialist_statuses (status_name) VALUES ('{row}')"
            self.cursor.execute(sql_script)
            print(f'Значение {row} добавлено!')

    def insert_specialists_data_in_db(self, specialists_data):
        """Добавление списка специалистов"""
        for row in specialists_data:
            sql_script = f"INSERT INTO specialists (specialist_name, fk_status_specialist) VALUES (" \
                         f"'{row['name']}'," \
                         f"(SELECT status_id FROM specialist_statuses WHERE status_name = '{row['status']}'))"
            self.cursor.execute(sql_script)
            print(f"Сотрудник {row['name']} добавлен!")

    def insert_group_services_data_in_db(self, group_services_data):
        """Добавление групп услуг"""
        for row in group_services_data:
            sql_script = f"INSERT INTO group_services (group_name) VALUES ('{row}')"
            self.cursor.execute(sql_script)
            print(f'Группа услуг {row} добавлена')

    def insert_tags_services_data_in_db(self, tags_services_data):
        """Добавление тег услуг"""
        for row in tags_services_data:
            sql_script = f"INSERT INTO tags_services (tag_name) VALUES ('{row}')"
            self.cursor.execute(sql_script)
            print(f'Тег {row} добавлен!')

    def insert_services_data_in_db(self, services_data):
        """Добавление списка услуг"""
        for row in services_data:
            sql_script = f"INSERT INTO services (service_name, fk_group_service) VALUES ('{row['name']}', " \
                         f"(SELECT group_id FROM group_services WHERE group_name = '{row['group']}'))"
            self.cursor.execute(sql_script)
            print(f"Услуга {row['name']} добавлена!")

    def insert_program_services_data_in_db(self, program_services_data):
        """Добавление программных услуг в БД"""
        for row in program_services_data:
            sql_script = f"INSERT INTO program_services (program_service_name, fk_tag_service) VALUES (" \
                         f"'{row['name']}', " \
                         f"(SELECT tag_id FROM tags_services WHERE tag_name = '{row['tag']}'))"
            self.cursor.execute(sql_script)
            print(f"Программная услуга {row['name']} добавлена!")

    def commit_bd(self):
        self.conn.commit()

    def close_bd(self):
        self.conn.close()


if __name__ == '__main__':
    excel_data = ExcelBase()
    bd = UpdateBase()

    # Создание таблиц БД
    bd.create_tablets()

    # Заполнение таблицы статусы специалистов
    status_data = excel_data.parse_status_specialist()
    bd.insert_specialist_status_data_in_db(status_data)

    # Заполнение таблицы списка специалистов
    specialists_data = excel_data.parse_specialists()
    bd.insert_specialists_data_in_db(specialists_data)

    # Заполнение таблицы списка групп услуг
    group_services_data = excel_data.parse_group_services()
    bd.insert_group_services_data_in_db(group_services_data)

    # Заполнение таблицы списка тегов услуг
    tags_services_data = excel_data.parse_tags_services()
    bd.insert_tags_services_data_in_db(tags_services_data)

    # Заполнение таблицы списка услуг
    services_data = excel_data.parse_services()
    bd.insert_services_data_in_db(services_data)

    # Заполнение таблицы списка программных услуг
    program_services_data = excel_data.parse_program_services()
    bd.insert_program_services_data_in_db(program_services_data)

    bd.commit_bd()
    bd.close_bd()
