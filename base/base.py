import psycopg2
from openpyxl import load_workbook
import config


class Base:
    def __init__(self):
        self.conn = psycopg2.connect(dbname=config.base_name, user=config.base_user,
                                     password=config.base_password, host=config.base_host)
        self.cursor = self.conn.cursor()

    def create_tablets(self):
        with open('CREATE_TABLES_SCRIPT.sql', 'r') as file:
            sql_script = file.read()
        self.cursor.execute(sql_script)

    def insert_specialist_status_data_in_db(self, status_data):
        for row in status_data:
            sql_script = f"INSERT INTO specialist_statuses (status_name) VALUES ('{row}')"
            self.cursor.execute(sql_script)
            print(f'Значение {row} добавлено!')

    def insert_specialists_data_in_db(self, specialists_data):
        for row in specialists_data:
            sql_script = f"INSERT INTO specialists (specialist_name, fk_status_specialist) VALUES (" \
                         f"'{row['name']}'," \
                         f"(SELECT status_id FROM specialist_statuses WHERE status_name = '{row['status']}'))"
            self.cursor.execute(sql_script)
            print(f"Сотрудник {row['name']} добавлен!")

    def insert_group_services_data_in_db(self, group_services_data):
        for row in group_services_data:
            sql_script = f"INSERT INTO group_services (group_name) VALUES ('{row}')"
            self.cursor.execute(sql_script)
            print(f'Группа услуг {row} добавлена')

    def insert_tags_services_data_in_db(self, tags_services_data):
        for row in tags_services_data:
            sql_script = f"INSERT INTO tags_services (tag_name) VALUES ('{row}')"
            self.cursor.execute(sql_script)
            print(f'Тег {row} добавлен!')


    def commit_bd(self):
        self.conn.commit()

    def close_bd(self):
        self.conn.close()


class ExcelBase:
    def __init__(self):
        self.wb = load_workbook('base_model.xlsx')

    def parse_status_specialist(self):
        sheet_status_specialist = self.wb['status_specialists']
        statuses_specialists = []
        rows = sheet_status_specialist.max_row
        for i in range(2, rows + 1):
            statuses_specialists.append(sheet_status_specialist.cell(row=i, column=1).value)
        return statuses_specialists

    def parse_specialists(self):
        sheet_specialists = self.wb['specialists']
        specialists = []
        rows = sheet_specialists.max_row
        for i in range(2, rows + 1):
            name = sheet_specialists.cell(row=i, column=1).value
            status = sheet_specialists.cell(row=i, column=2).value
            specialists.append({'name': name, 'status': status})
        return specialists

    def parse_group_services(self):
        sheet_group_services = self.wb['groups']
        groups_services = []
        rows = sheet_group_services.max_row
        for i in range(2, rows + 1):
            groups_services.append(sheet_group_services.cell(row=i, column=1).value)
        return groups_services

    def parse_tags_services(self):
        sheet_group_tags_services = self.wb['tegs']
        tags_services = []
        rows = sheet_group_tags_services.max_row
        for i in range(2, rows + 1):
            tags_services.append(sheet_group_tags_services.cell(row=i, column=1).value)
        return tags_services


if __name__ == '__main__':
    bd = Base()
    excel_data = ExcelBase()

    # Заполнение таблицы статусы специалистов
    # status_data = excel_data.parse_status_specialist()
    # bd.insert_specialist_status_data_in_db(status_data)

    # Заполнение таблицы списка специалистов
    # specialists_data = excel_data.parse_specialists()
    # bd.insert_specialists_data_in_db(specialists_data)

    # Заполнение таблицы списка групп услуг
    # group_services_data = excel_data.parse_group_services()
    # bd.insert_group_services_data_in_db(group_services_data)

    # Заполнение таблицы списка тегов услуг
    tags_services_data = excel_data.parse_tags_services()
    bd.insert_tags_services_data_in_db(tags_services_data)

    bd.commit_bd()
    bd.close_bd()
